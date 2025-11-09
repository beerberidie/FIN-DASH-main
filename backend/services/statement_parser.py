"""
Statement Parser Service
Parses bank statements from various formats (CSV, Excel, PDF, OFX, QFX)
"""

import csv
import io
import re
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
import pdfplumber
import xlrd
from openpyxl import load_workbook
import ofxparse
from fuzzywuzzy import fuzz


class StatementParser:
    """Parse bank statements from various file formats."""
    
    # Common date formats
    DATE_FORMATS = [
        '%Y-%m-%d',
        '%d/%m/%Y',
        '%m/%d/%Y',
        '%d-%m-%Y',
        '%Y/%m/%d',
        '%d.%m.%Y',
        '%d %b %Y',
        '%d %B %Y',
        '%b %d, %Y',
        '%B %d, %Y',
    ]
    
    # Common column name patterns
    COLUMN_PATTERNS = {
        'date': ['date', 'transaction date', 'posting date', 'value date', 'trans date'],
        'description': ['description', 'details', 'narrative', 'particulars', 'memo', 'reference'],
        'amount': ['amount', 'value', 'transaction amount', 'debit', 'credit'],
        'debit': ['debit', 'withdrawal', 'out', 'payment'],
        'credit': ['credit', 'deposit', 'in', 'receipt'],
        'balance': ['balance', 'running balance', 'closing balance', 'available balance'],
    }
    
    @staticmethod
    def detect_file_type(file_path: Path) -> str:
        """Detect file type from extension."""
        suffix = file_path.suffix.lower()
        if suffix == '.csv':
            return 'csv'
        elif suffix in ['.xls', '.xlsx']:
            return 'excel'
        elif suffix == '.pdf':
            return 'pdf'
        elif suffix in ['.ofx', '.qfx']:
            return 'ofx'
        else:
            raise ValueError(f"Unsupported file type: {suffix}")
    
    @staticmethod
    def parse_date(date_str: str) -> Optional[str]:
        """Parse date string to ISO format."""
        if not date_str or not isinstance(date_str, str):
            return None
        
        date_str = date_str.strip()
        
        for fmt in StatementParser.DATE_FORMATS:
            try:
                dt = datetime.strptime(date_str, fmt)
                return dt.strftime('%Y-%m-%d')
            except ValueError:
                continue
        
        return None
    
    @staticmethod
    def parse_amount(amount_str: str) -> Optional[float]:
        """Parse amount string to float."""
        if not amount_str:
            return None
        
        # Convert to string if not already
        amount_str = str(amount_str).strip()
        
        # Remove currency symbols and spaces
        amount_str = re.sub(r'[R$€£¥\s]', '', amount_str)
        
        # Handle parentheses for negative amounts
        if amount_str.startswith('(') and amount_str.endswith(')'):
            amount_str = '-' + amount_str[1:-1]
        
        # Remove thousands separators
        amount_str = amount_str.replace(',', '')
        
        try:
            return float(amount_str)
        except ValueError:
            return None
    
    @staticmethod
    def detect_column_mapping(headers: List[str]) -> Dict[str, int]:
        """Detect column mapping from headers."""
        headers_lower = [h.lower().strip() for h in headers]
        mapping = {}
        
        for field, patterns in StatementParser.COLUMN_PATTERNS.items():
            for i, header in enumerate(headers_lower):
                for pattern in patterns:
                    if fuzz.ratio(header, pattern) > 80:
                        mapping[field] = i
                        break
                if field in mapping:
                    break
        
        return mapping
    
    @staticmethod
    def parse_csv(file_path: Path) -> List[Dict[str, Any]]:
        """Parse CSV bank statement."""
        transactions = []
        
        # Try different encodings
        encodings = ['utf-8', 'latin-1', 'cp1252']
        content = None
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    content = f.read()
                break
            except UnicodeDecodeError:
                continue
        
        if not content:
            raise ValueError("Could not decode CSV file")
        
        # Detect delimiter
        sniffer = csv.Sniffer()
        delimiter = sniffer.sniff(content[:1024]).delimiter
        
        # Parse CSV
        reader = csv.reader(io.StringIO(content), delimiter=delimiter)
        rows = list(reader)
        
        if not rows:
            return transactions
        
        # Detect header row
        headers = rows[0]
        column_mapping = StatementParser.detect_column_mapping(headers)
        
        # Parse transactions
        for row in rows[1:]:
            if not row or len(row) < 2:
                continue
            
            transaction = {}
            
            # Extract date
            if 'date' in column_mapping:
                date_str = row[column_mapping['date']]
                transaction['date'] = StatementParser.parse_date(date_str)
            
            # Extract description
            if 'description' in column_mapping:
                transaction['description'] = row[column_mapping['description']].strip()
            
            # Extract amount
            amount = None
            if 'amount' in column_mapping:
                amount = StatementParser.parse_amount(row[column_mapping['amount']])
            elif 'debit' in column_mapping and 'credit' in column_mapping:
                debit = StatementParser.parse_amount(row[column_mapping['debit']])
                credit = StatementParser.parse_amount(row[column_mapping['credit']])
                if debit:
                    amount = -abs(debit)
                elif credit:
                    amount = abs(credit)
            
            if amount is not None:
                transaction['amount'] = amount
            
            # Extract balance
            if 'balance' in column_mapping:
                balance = StatementParser.parse_amount(row[column_mapping['balance']])
                if balance is not None:
                    transaction['balance'] = balance
            
            # Only add if we have minimum required fields
            if transaction.get('date') and transaction.get('description') and 'amount' in transaction:
                transactions.append(transaction)
        
        return transactions
    
    @staticmethod
    def parse_excel(file_path: Path) -> List[Dict[str, Any]]:
        """Parse Excel bank statement."""
        transactions = []
        
        # Try openpyxl first (for .xlsx)
        if file_path.suffix.lower() == '.xlsx':
            try:
                wb = load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.values)
                wb.close()
            except Exception:
                # Fall back to xlrd
                rows = StatementParser._parse_excel_xlrd(file_path)
        else:
            # Use xlrd for .xls
            rows = StatementParser._parse_excel_xlrd(file_path)
        
        if not rows:
            return transactions
        
        # Detect header row
        headers = [str(cell) if cell else '' for cell in rows[0]]
        column_mapping = StatementParser.detect_column_mapping(headers)
        
        # Parse transactions
        for row in rows[1:]:
            if not row or all(cell is None or cell == '' for cell in row):
                continue
            
            transaction = {}
            
            # Extract date
            if 'date' in column_mapping and len(row) > column_mapping['date']:
                date_cell = row[column_mapping['date']]
                if isinstance(date_cell, datetime):
                    transaction['date'] = date_cell.strftime('%Y-%m-%d')
                else:
                    transaction['date'] = StatementParser.parse_date(str(date_cell))
            
            # Extract description
            if 'description' in column_mapping and len(row) > column_mapping['description']:
                desc = row[column_mapping['description']]
                if desc:
                    transaction['description'] = str(desc).strip()
            
            # Extract amount
            amount = None
            if 'amount' in column_mapping and len(row) > column_mapping['amount']:
                amount = StatementParser.parse_amount(str(row[column_mapping['amount']]))
            elif 'debit' in column_mapping and 'credit' in column_mapping:
                debit_val = row[column_mapping['debit']] if len(row) > column_mapping['debit'] else None
                credit_val = row[column_mapping['credit']] if len(row) > column_mapping['credit'] else None
                
                debit = StatementParser.parse_amount(str(debit_val)) if debit_val else None
                credit = StatementParser.parse_amount(str(credit_val)) if credit_val else None
                
                if debit:
                    amount = -abs(debit)
                elif credit:
                    amount = abs(credit)
            
            if amount is not None:
                transaction['amount'] = amount
            
            # Extract balance
            if 'balance' in column_mapping and len(row) > column_mapping['balance']:
                balance = StatementParser.parse_amount(str(row[column_mapping['balance']]))
                if balance is not None:
                    transaction['balance'] = balance
            
            # Only add if we have minimum required fields
            if transaction.get('date') and transaction.get('description') and 'amount' in transaction:
                transactions.append(transaction)
        
        return transactions
    
    @staticmethod
    def _parse_excel_xlrd(file_path: Path) -> List[List[Any]]:
        """Parse Excel file using xlrd."""
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)
        
        rows = []
        for row_idx in range(sheet.nrows):
            row = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    date_tuple = xlrd.xldate_as_tuple(cell.value, workbook.datemode)
                    row.append(datetime(*date_tuple))
                else:
                    row.append(cell.value)
            rows.append(row)
        
        return rows

    @staticmethod
    def parse_pdf(file_path: Path) -> List[Dict[str, Any]]:
        """Parse PDF bank statement."""
        transactions = []

        with pdfplumber.open(file_path) as pdf:
            all_text = ""
            for page in pdf.pages:
                all_text += page.extract_text() + "\n"

        # Try to extract transactions using regex patterns
        # Pattern 1: Date Description Amount Balance
        pattern1 = r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+(.+?)\s+([-]?[\d,]+\.\d{2})\s+([\d,]+\.\d{2})'

        # Pattern 2: Date Description Debit Credit Balance
        pattern2 = r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+(.+?)\s+([-]?[\d,]+\.\d{2})?\s+([-]?[\d,]+\.\d{2})?\s+([\d,]+\.\d{2})'

        lines = all_text.split('\n')

        for line in lines:
            # Try pattern 1
            match = re.search(pattern1, line)
            if match:
                date_str, description, amount_str, balance_str = match.groups()
                date = StatementParser.parse_date(date_str)
                amount = StatementParser.parse_amount(amount_str)
                balance = StatementParser.parse_amount(balance_str)

                if date and description and amount is not None:
                    transaction = {
                        'date': date,
                        'description': description.strip(),
                        'amount': amount,
                    }
                    if balance is not None:
                        transaction['balance'] = balance
                    transactions.append(transaction)
                continue

            # Try pattern 2
            match = re.search(pattern2, line)
            if match:
                date_str, description, debit_str, credit_str, balance_str = match.groups()
                date = StatementParser.parse_date(date_str)

                amount = None
                if debit_str:
                    debit = StatementParser.parse_amount(debit_str)
                    if debit:
                        amount = -abs(debit)
                if credit_str and amount is None:
                    credit = StatementParser.parse_amount(credit_str)
                    if credit:
                        amount = abs(credit)

                balance = StatementParser.parse_amount(balance_str)

                if date and description and amount is not None:
                    transaction = {
                        'date': date,
                        'description': description.strip(),
                        'amount': amount,
                    }
                    if balance is not None:
                        transaction['balance'] = balance
                    transactions.append(transaction)

        return transactions

    @staticmethod
    def parse_ofx(file_path: Path) -> List[Dict[str, Any]]:
        """Parse OFX/QFX bank statement."""
        transactions = []

        with open(file_path, 'rb') as f:
            ofx = ofxparse.OfxParser.parse(f)

        # Get the first account
        if not ofx.accounts:
            return transactions

        account = ofx.accounts[0]

        for txn in account.statement.transactions:
            transaction = {
                'date': txn.date.strftime('%Y-%m-%d'),
                'description': txn.memo or txn.payee or 'Unknown',
                'amount': float(txn.amount),
            }

            # Add transaction ID if available
            if txn.id:
                transaction['external_id'] = txn.id

            transactions.append(transaction)

        return transactions

    @staticmethod
    def parse_file(file_path: Path) -> Tuple[List[Dict[str, Any]], str]:
        """
        Parse bank statement file.
        Returns tuple of (transactions, file_type)
        """
        file_type = StatementParser.detect_file_type(file_path)

        if file_type == 'csv':
            transactions = StatementParser.parse_csv(file_path)
        elif file_type == 'excel':
            transactions = StatementParser.parse_excel(file_path)
        elif file_type == 'pdf':
            transactions = StatementParser.parse_pdf(file_path)
        elif file_type == 'ofx':
            transactions = StatementParser.parse_ofx(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")

        return transactions, file_type

