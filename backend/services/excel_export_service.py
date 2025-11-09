"""Excel and CSV export service."""
import os
import csv
from datetime import datetime, date as date_type
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.export import ExportConfig
from services.csv_manager import csv_manager
from services.calculator import calculator
from services.debt_service import debt_service
from services.portfolio_service import portfolio_service
from utils.dates import get_current_month


class ExcelExportService:
    """Service for generating Excel and CSV exports."""
    
    def __init__(self):
        """Initialize Excel export service."""
        self.exports_dir = "exports"
        self._ensure_exports_dir()
    
    def _ensure_exports_dir(self):
        """Ensure exports directory exists."""
        if not os.path.exists(self.exports_dir):
            os.makedirs(self.exports_dir)
    
    def _format_currency(self, amount: float, config: ExportConfig) -> str:
        """Format currency value."""
        formatted = f"{amount:{config.number_format}}"
        return f"{config.base_currency} {formatted}"
    
    def _apply_header_style(self, ws, row: int, max_col: int):
        """Apply header styling to a row."""
        header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        header_font = Font(bold=True, size=11)
        
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def _apply_border(self, ws, start_row: int, end_row: int, max_col: int):
        """Apply borders to a range."""
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).border = thin_border
    
    def _auto_size_columns(self, ws, max_col: int):
        """Auto-size columns based on content."""
        for col in range(1, max_col + 1):
            max_length = 0
            column = get_column_letter(col)
            
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    def export_transactions_excel(
        self,
        start_date: Optional[date_type],
        end_date: Optional[date_type],
        account_id: Optional[str],
        category_id: Optional[str],
        transaction_type: Optional[str],
        config: ExportConfig
    ) -> str:
        """
        Export transactions to Excel.
        
        Returns:
            File path of generated Excel file
        """
        # Load data
        transactions = csv_manager.read_csv("transactions.csv")
        accounts = csv_manager.read_csv("accounts.csv")
        categories = csv_manager.read_csv("categories.csv")
        
        # Create lookup maps
        account_map = {acc['id']: acc['name'] for acc in accounts}
        category_map = {cat['id']: cat['name'] for cat in categories}
        
        # Filter transactions
        filtered = []
        for txn in transactions:
            # Date filter
            if start_date and txn.get('date', '') < str(start_date):
                continue
            if end_date and txn.get('date', '') > str(end_date):
                continue
            
            # Account filter
            if account_id and txn.get('account_id') != account_id:
                continue
            
            # Category filter
            if category_id and txn.get('category_id') != category_id:
                continue
            
            # Type filter
            if transaction_type and txn.get('type') != transaction_type:
                continue
            
            filtered.append(txn)
        
        # Sort by date descending
        filtered.sort(key=lambda x: x.get('date', ''), reverse=True)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Title
        ws['A1'] = "Transaction Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        # Date range
        date_range = ""
        if start_date and end_date:
            date_range = f"{start_date} to {end_date}"
        elif start_date:
            date_range = f"From {start_date}"
        elif end_date:
            date_range = f"Until {end_date}"
        
        if date_range:
            ws['A2'] = date_range
            ws['A2'].font = Font(size=10, color="6B7280")
            ws.merge_cells('A2:F2')
        
        # Summary
        total_income = sum(float(t.get('amount', 0)) for t in filtered if t.get('type') == 'income')
        total_expenses = sum(abs(float(t.get('amount', 0))) for t in filtered if t.get('type') == 'expense')
        net = total_income - total_expenses
        
        summary_row = 4
        ws[f'A{summary_row}'] = "Total Income:"
        ws[f'B{summary_row}'] = self._format_currency(total_income, config)
        ws[f'A{summary_row}'].font = Font(bold=True)
        
        ws[f'A{summary_row + 1}'] = "Total Expenses:"
        ws[f'B{summary_row + 1}'] = self._format_currency(total_expenses, config)
        ws[f'A{summary_row + 1}'].font = Font(bold=True)
        
        ws[f'A{summary_row + 2}'] = "Net:"
        ws[f'B{summary_row + 2}'] = self._format_currency(net, config)
        ws[f'A{summary_row + 2}'].font = Font(bold=True)
        
        # Headers
        header_row = summary_row + 4
        headers = ['Date', 'Description', 'Category', 'Account', 'Type', 'Amount', 'Currency', 'Notes']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col, value=header)
        
        self._apply_header_style(ws, header_row, len(headers))
        
        # Data
        data_start_row = header_row + 1
        for row_idx, txn in enumerate(filtered, start=data_start_row):
            amount = float(txn.get('amount', 0))
            
            ws.cell(row=row_idx, column=1, value=txn.get('date', ''))
            ws.cell(row=row_idx, column=2, value=txn.get('description', ''))
            ws.cell(row=row_idx, column=3, value=category_map.get(txn.get('category_id', ''), 'Unknown'))
            ws.cell(row=row_idx, column=4, value=account_map.get(txn.get('account_id', ''), 'Unknown'))
            ws.cell(row=row_idx, column=5, value=txn.get('type', ''))
            ws.cell(row=row_idx, column=6, value=abs(amount))
            ws.cell(row=row_idx, column=7, value=txn.get('currency', config.base_currency))
            ws.cell(row=row_idx, column=8, value=txn.get('notes', ''))
        
        # Apply borders
        if filtered:
            self._apply_border(ws, header_row, data_start_row + len(filtered) - 1, len(headers))
        
        # Auto-size columns
        self._auto_size_columns(ws, len(headers))
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"transactions_{timestamp}.xlsx"
        filepath = os.path.join(self.exports_dir, filename)
        
        # Save workbook
        wb.save(filepath)
        
        return filepath
    
    def export_transactions_csv(
        self,
        start_date: Optional[date_type],
        end_date: Optional[date_type],
        account_id: Optional[str],
        category_id: Optional[str],
        transaction_type: Optional[str],
        config: ExportConfig
    ) -> str:
        """
        Export transactions to CSV.
        
        Returns:
            File path of generated CSV file
        """
        # Load data
        transactions = csv_manager.read_csv("transactions.csv")
        accounts = csv_manager.read_csv("accounts.csv")
        categories = csv_manager.read_csv("categories.csv")
        
        # Create lookup maps
        account_map = {acc['id']: acc['name'] for acc in accounts}
        category_map = {cat['id']: cat['name'] for cat in categories}
        
        # Filter transactions
        filtered = []
        for txn in transactions:
            # Date filter
            if start_date and txn.get('date', '') < str(start_date):
                continue
            if end_date and txn.get('date', '') > str(end_date):
                continue
            
            # Account filter
            if account_id and txn.get('account_id') != account_id:
                continue
            
            # Category filter
            if category_id and txn.get('category_id') != category_id:
                continue
            
            # Type filter
            if transaction_type and txn.get('type') != transaction_type:
                continue
            
            filtered.append(txn)
        
        # Sort by date descending
        filtered.sort(key=lambda x: x.get('date', ''), reverse=True)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"transactions_{timestamp}.csv"
        filepath = os.path.join(self.exports_dir, filename)
        
        # Write CSV
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            fieldnames = ['Date', 'Description', 'Category', 'Account', 'Type', 'Amount', 'Currency', 'Notes', 'Tags']
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            
            writer.writeheader()
            
            for txn in filtered:
                writer.writerow({
                    'Date': txn.get('date', ''),
                    'Description': txn.get('description', ''),
                    'Category': category_map.get(txn.get('category_id', ''), 'Unknown'),
                    'Account': account_map.get(txn.get('account_id', ''), 'Unknown'),
                    'Type': txn.get('type', ''),
                    'Amount': txn.get('amount', ''),
                    'Currency': txn.get('currency', config.base_currency),
                    'Notes': txn.get('notes', ''),
                    'Tags': txn.get('tags', '')
                })
        
        return filepath
    
    def export_investment_portfolio_excel(
        self,
        include_transactions: bool,
        investment_type: Optional[str],
        config: ExportConfig
    ) -> str:
        """
        Export investment portfolio to Excel.
        
        Returns:
            File path of generated Excel file
        """
        from services.investment_service import investment_service
        
        # Load investments
        investments = investment_service.list_investments(type_filter=investment_type)
        
        # Get portfolio summary
        portfolio_summary = portfolio_service.get_portfolio_summary(config.base_currency)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Portfolio"
        
        # Title
        ws['A1'] = "Investment Portfolio"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        
        # Summary
        summary_row = 3
        ws[f'A{summary_row}'] = "Total Investments:"
        ws[f'B{summary_row}'] = portfolio_summary.total_investments
        ws[f'A{summary_row}'].font = Font(bold=True)
        
        ws[f'A{summary_row + 1}'] = "Total Value:"
        ws[f'B{summary_row + 1}'] = self._format_currency(portfolio_summary.total_value, config)
        ws[f'A{summary_row + 1}'].font = Font(bold=True)
        
        ws[f'A{summary_row + 2}'] = "Total Cost:"
        ws[f'B{summary_row + 2}'] = self._format_currency(portfolio_summary.total_cost, config)
        ws[f'A{summary_row + 2}'].font = Font(bold=True)
        
        ws[f'A{summary_row + 3}'] = "Profit/Loss:"
        ws[f'B{summary_row + 3}'] = self._format_currency(portfolio_summary.total_profit_loss, config)
        ws[f'A{summary_row + 3}'].font = Font(bold=True)
        
        ws[f'A{summary_row + 4}'] = "Return:"
        ws[f'B{summary_row + 4}'] = f"{portfolio_summary.total_profit_loss_percentage:.2f}%"
        ws[f'A{summary_row + 4}'].font = Font(bold=True)
        
        # Headers
        header_row = summary_row + 6
        headers = ['Symbol', 'Name', 'Type', 'Quantity', 'Avg Cost', 'Current Price', 'Total Value', 'P/L %']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col, value=header)
        
        self._apply_header_style(ws, header_row, len(headers))
        
        # Data
        data_start_row = header_row + 1
        for row_idx, inv in enumerate(investments, start=data_start_row):
            perf = investment_service.get_investment_performance(inv.id)
            
            ws.cell(row=row_idx, column=1, value=inv.symbol)
            ws.cell(row=row_idx, column=2, value=inv.name)
            ws.cell(row=row_idx, column=3, value=inv.type)
            ws.cell(row=row_idx, column=4, value=perf.quantity)
            ws.cell(row=row_idx, column=5, value=perf.average_cost)
            ws.cell(row=row_idx, column=6, value=perf.current_price)
            ws.cell(row=row_idx, column=7, value=perf.current_value)
            ws.cell(row=row_idx, column=8, value=f"{perf.profit_loss_percentage:.2f}%")
        
        # Apply borders
        if investments:
            self._apply_border(ws, header_row, data_start_row + len(investments) - 1, len(headers))
        
        # Auto-size columns
        self._auto_size_columns(ws, len(headers))
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"investment_portfolio_{timestamp}.xlsx"
        filepath = os.path.join(self.exports_dir, filename)
        
        # Save workbook
        wb.save(filepath)
        
        return filepath


# Singleton instance
excel_export_service = ExcelExportService()

